import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import mysql.connector
from database import Database

# Veritabanı bağlantısı
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "1234",  # Şifrenizi ekleyin
    "database": "app"  # Veritabanı adınızı yazın
}

db = Database(**DB_CONFIG)

# Yeni bir Excel çalışma kitabı oluştur
wb = Workbook()
ws = wb.active
ws.title = "Sayfa1"

# Başlık satırlarını ekleyelim
headers = ["Bölüm", "", "BÖLÜM ADI", "", "", ""]
sub_headers = ["", "", "1. Sınıf", "2. Sınıf", "3. Sınıf", "4. Sınıf"]
ws.append(headers)
ws.append(sub_headers)

# Gün ve saat tablosu
days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
hours = ["09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00", "16:00-17:00"]

schedule = []
for day in days:
    schedule.append([day] + [hours[0]] + ["", "", "", ""])
    for hour in hours[1:]:
        schedule.append([""] + [hour] + ["", "", "", ""])

for row in schedule:
    ws.append(row)

# Hücre genişliklerini ve satır yüksekliğini ayarlayalım
column_widths = [50, 60, 70, 70, 70, 70]  # Sütun genişliklerini arttırdık
for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# Satır yüksekliğini arttırıyoruz
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row[0].row].height = 50  # Satır yüksekliği

# Hücrelerin kenarlıklarını ayarlayalım
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border

# Başlıkların biçimlendirilmesi
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
center_alignment = Alignment(horizontal="center", vertical="center")

for col in range(1, 7):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

for col in range(3, 7):
    cell = ws.cell(row=2, column=col)
    cell.alignment = center_alignment
    cell.border = thin_border

# Dersleri veritabanından çekme
query = """
SELECT d.ders_adi, d.haftalik_saat, b.bolum_adi, d.donem, k.ad_soyad, dl.derslik_id
FROM Dersler d
JOIN Bolum b ON d.bolum_kodu = b.bolum_kodu
JOIN Kullanici k ON d.ogretim_uyesi_id = k.id
JOIN Derslik dl ON d.derslik_id = dl.derslik_id
WHERE b.bolum_adi IN ('Yazılım Mühendisliği', 'Ortak')
"""
dersler = db.fetch_all(query)

# Dersleri Excel'e yerleştirme
row_index = 3  # 3. satırdan itibaren dersleri yerleştiriyoruz

# Her dersin sınıfına ve dönemine göre sütunlar
donem_to_column = {
    2: 3,  # 1. sınıf -> 2. dönem -> 3. sütun
    4: 4,  # 2. sınıf -> 4. dönem -> 4. sütun
    6: 5,  # 3. sınıf -> 6. dönem -> 5. sütun
    8: 6   # 4. sınıf -> 8. dönem -> 6. sütun
}

# Haftalık saat sayısına göre derslerin satırlara yerleştirilmesi
# Dersin yerleşeceği hücreyi daha düzgün bir şekilde yerleştiriyoruz.
for ders in dersler:
    ders_adi, haftalik_saat, bolum_adi, donem, ogretim_uyesi, derslik_id = ders
    column = donem_to_column.get(donem)
    
    if column:
        for i in range(haftalik_saat):
            # Öğretim üyesi ve derslik bilgisi de ekleniyor
            value = f"{ders_adi}\nÖğretim Üyesi: {ogretim_uyesi}\nDerslik: {derslik_id}"
            ws.cell(row=row_index, column=column, value=value)
            row_index += 1
            if row_index > ws.max_row:  # Saatler dolarsa başa sarıyoruz
                row_index = 3

# Excel dosyasını kaydetme
output_file = "output_yazilim.xlsx"
wb.save(output_file)

print(f"Excel dosyası '{output_file}' olarak derslerle dolduruldu.")
