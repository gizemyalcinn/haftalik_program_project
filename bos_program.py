import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Yeni bir Excel çalışma kitabı oluştur
wb = Workbook()
ws = wb.active
ws.title = "Sayfa1"

# Başlık satırlarını ekleyelim
headers = ["Bölüm", "", "BÖLÜM ADI", "", "", ""]
sub_headers = ["", "", "1. Sınıf", "2. Sınıf", "3. Sınıf", "4. Sınıf"]
ws.append(headers)
ws.append(sub_headers)

# Gün ve saat tablosu
schedule = []
days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
hours = ["09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00", "16:00-17:00"]

for day in days:
    schedule.append([day] + [hours[0]] + ["", "", "", ""])
    for hour in hours[1:]:
        schedule.append([""] + [hour] + ["", "", "", ""])

for row in schedule:
    ws.append(row)

# Hücre genişliklerini ve satır yüksekliğini ayarlayalım
column_widths = [40, 50, 60, 60, 60, 60]  # Sütun genişliklerini arttırdık
for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# Satır yüksekliğini arttırıyoruz
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row[0].row].height = 40  # Satır yüksekliği

# Hücrelerin kenarlıklarını ayarlayalım
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border

# Başlıkların biçimlendirilmesi
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
center_alignment = Alignment(horizontal="center", vertical="center")

for col in range(1, 7):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

for col in range(3, 7):
    cell = ws.cell(row=2, column=col)
    cell.alignment = center_alignment
    cell.border = thin_border

# Excel dosyasını kaydetme
output_file = "output.xlsx"
wb.save(output_file)

print(f"Excel dosyası '{output_file}' olarak birebir oluşturuldu.")
